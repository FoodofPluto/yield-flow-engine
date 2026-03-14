# yf_ingest.py
import re
import sys
import argparse
from pathlib import Path
from datetime import datetime, timezone
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    print("Missing dependency: openpyxl. Install with:  pip install openpyxl  (or poetry add openpyxl)")
    sys.exit(1)

# ---------- Timezone handling (robust on Windows) ----------
def resolve_tz(key: str):
    try:
        return ZoneInfo(key)
    except ZoneInfoNotFoundError:
        try:
            import tzdata  # noqa: F401
            return ZoneInfo(key)
        except Exception:
            try:
                return datetime.now().astimezone().tzinfo
            except Exception:
                return timezone.utc

LOCAL_TZ = resolve_tz("America/New_York")
RUN_TS = datetime.now(LOCAL_TZ)
TZ_LABEL = RUN_TS.tzname() or "Local"
RUN_TS_STR = RUN_TS.strftime(f"%Y-%m-%d %H:%M:%S {TZ_LABEL}")

# ---------- Parsing ----------
ROW_RX = re.compile(
    r"""
    ^(?P<Name>.+?)               # protocol name (lazy)
    \s+                          # spaces
    (?P<APY>[-+]?\d[\d,]*\.?\d*) # APY number like 9,229.01
    \s+                          # spaces
    (?P<Source>\S+)              # defillama
    \s+                          # spaces
    (?P<TVL>[-+]?\d[\d,]*)       # TVL like 19,559
    \s+                          # spaces
    (?P<Chain>\S+)               # Arbitrum / Base / etc
    \s*$
    """,
    re.VERBOSE,
)
HEADER_RX = re.compile(r"^\s*Name\s+APY%.*TVL.*Chain\s*$", re.IGNORECASE)
RULER_RX = re.compile(r"^-{3,}")

def parse_rows(text: str):
    rows = []
    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue
        if HEADER_RX.search(line) or RULER_RX.search(line):
            continue
        m = ROW_RX.match(line)
        if not m:
            # Uncomment to debug:
            # print(f"SKIP: {line}")
            continue
        d = m.groupdict()
        # normalize fields
        apy = d["APY"].replace(",", "").rstrip("%")
        tvl = d["TVL"].replace(",", "")
        rows.append({
            "Name": d["Name"],
            "APY": float(apy) if apy else None,
            "Source": d["Source"],
            "TVL": float(tvl) if tvl else None,
            "Chain": d["Chain"],
        })
    return rows

# ---------- Workbook writing ----------
EXPECTED_HEADERS = ["Name", "APY", "Source", "TVL", "Chain", "Timestamp"]

def ensure_headers(ws):
    # If empty sheet or blank header row, write fresh headers
    if ws.max_row == 0 or (ws.max_row == 1 and all(c.value in (None, "") for c in ws[1])):
        ws.append(EXPECTED_HEADERS)
        return

    # Current headers
    current = [c.value.strip() if isinstance(c.value, str) else c.value for c in ws[1]]
    # Normalize common variants
    fixes = {"APY%": "APY", "TVL (USD)": "TVL"}
    current = [fixes.get(h, h) for h in current]

    # If invalid header row, replace it
    if not current or any(h in (None, "") for h in current):
        ws.delete_rows(1, 1)
        ws.append(EXPECTED_HEADERS)
        return

    # Append any missing headers at the end
    for h in EXPECTED_HEADERS:
        if h not in current:
            ws.cell(row=1, column=len(current) + 1, value=h)
            current.append(h)

def write_to_workbook(rows, workbook_path: Path, sheet_name="Ingest"):
    if workbook_path.exists():
        wb = load_workbook(workbook_path)
    else:
        wb = Workbook()

    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(title=sheet_name)
    ensure_headers(ws)

    # Append rows (header order) + run timestamp
    for r in rows:
        ordered = [
            r.get("Name"),
            r.get("APY"),
            r.get("Source"),
            r.get("TVL"),
            r.get("Chain"),
            RUN_TS_STR,
        ]
        ws.append(ordered)

    # Remove default empty sheet if present
    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1 and all(c.value is None for c in wb["Sheet"][1]):
        del wb["Sheet"]

    wb.save(workbook_path)

# ---------- CLI ----------
def main():
    p = argparse.ArgumentParser()
    p.add_argument("--input", required=True, help="Path to last-scan.txt (raw table output).")
    p.add_argument("--workbook", required=True, help="Path to Excel workbook (will be created if missing).")
    p.add_argument("--sheet", default="Ingest", help="Target sheet name (default: Ingest).")
    args = p.parse_args()

    txt = Path(args.input).read_text(encoding="utf-8", errors="replace")
    rows = parse_rows(txt)

    if not rows:
        print("No rows parsed. Make sure the input is the raw engine table output.")
        sys.exit(2)

    write_to_workbook(rows, Path(args.workbook), sheet_name=args.sheet)
    print(f"OK: wrote {len(rows)} rows to {args.workbook} [{args.sheet}]")

if __name__ == "__main__":
    main()
